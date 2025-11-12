"""
Google Maps Business Scraper - Robust Version
WARNING: This script is for educational purposes only.
Ensure you comply with Google's ToS and applicable laws.
"""

import asyncio
import json
import re
from playwright.async_api import async_playwright
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

class GoogleMapsScraper:
    def __init__(self):
        self.results = []
    
    async def scrape_businesses(self, search_query, max_results=None):
        """
        Scrape business information from Google Maps
        
        Args:
            search_query: Search term (e.g., "Freight Broker Doral")
            max_results: Maximum number of businesses to scrape (None = unlimited, scrape until end)
        """
        async with async_playwright() as p:
            # Launch browser (headless=False to see what's happening)
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            )
            page = await context.new_page()
            
            try:
                print(f"Searching for: {search_query}")
                
                # Navigate to Google Maps
                await page.goto('https://www.google.com/maps', timeout=60000)
                await page.wait_for_timeout(2000)
                
                # Search for the query
                search_box = await page.wait_for_selector('input#searchboxinput', timeout=10000)
                await search_box.fill(search_query)
                await page.keyboard.press('Enter')
                
                # Wait for results to load
                await page.wait_for_timeout(5000)
                
                # Scroll to load ALL results with robust handling
                print("Loading all results by scrolling...")
                await self._scroll_results_robust(page, max_results)
                
                # Get all business URLs (not elements, to avoid DOM detachment)
                print("\nCollecting business URLs...")
                business_urls = await self._get_all_business_urls(page)
                
                total_to_scrape = len(business_urls) if max_results is None else min(len(business_urls), max_results)
                print(f"\n✓ Found {len(business_urls)} unique businesses")
                print(f"Will scrape {total_to_scrape} businesses\n")
                
                # Scrape each business by navigating to URL
                businesses_to_scrape = business_urls if max_results is None else business_urls[:max_results]
                
                for i, url in enumerate(businesses_to_scrape):
                    try:
                        print(f"Scraping business {i+1}/{total_to_scrape}", end='')
                        
                        # Navigate directly to the business URL
                        await page.goto(url, timeout=30000)
                        await page.wait_for_timeout(2500)
                        
                        business_data = await self._extract_business_data(page)
                        if business_data and business_data.get('name') != 'not_available':
                            business_data['url'] = url  # Add the URL to data
                            self.results.append(business_data)
                            print(f" ✓ {business_data.get('name', 'Unknown')}")
                        else:
                            print(f" ⚠ Could not extract data")
                        
                    except Exception as e:
                        print(f" ✗ Error: {str(e)[:50]}")
                        continue
                
            except Exception as e:
                print(f"Error during scraping: {e}")
            
            finally:
                await browser.close()
        
        return self.results
    
    async def _scroll_results_robust(self, page, max_results):
        """
        ROBUST scrolling - keeps going until we see the end message (Google only shows it when TRULY done)
        """
        results_panel = await page.query_selector('div[role="feed"]')
        if not results_panel:
            print("⚠ Could not find results panel")
            return
        
        print("Scrolling until end message appears...")
        scroll_count = 0
        previous_count = 0
        no_change_count = 0
        
        while True:
            # Scroll down in the results panel
            await results_panel.evaluate('el => el.scrollBy(0, 1500)')
            scroll_count += 1
            
            # Initial wait
            await page.wait_for_timeout(1500)
            
            # Wait for any loading to finish
            await self._wait_for_loading_complete(page)
            
            # Count current results
            current_results = await page.query_selector_all('a[href*="/maps/place/"]')
            current_count = len(current_results)
            
            # Check for the ACTUAL end message - if we see this, we're DONE
            end_message = await page.query_selector('span:has-text("You\'ve reached the end of the list.")')
            
            if end_message:
                print(f"\n✓ Found end message! Total results: {current_count}")
                break
            
            # Display progress
            print(f"Scroll #{scroll_count}: {current_count} results", end='\r')
            
            # Check if we've hit max_results limit (if set)
            if max_results is not None and current_count >= max_results:
                print(f"\n✓ Reached target of {max_results} results!")
                break
            
            # Track stalls (no new results for many scrolls - backup safety)
            if current_count == previous_count:
                no_change_count += 1
                # Only stop from stall if we've tried MANY times (backup measure)
                if no_change_count >= 10:
                    print(f"\n⚠ Stalled at {current_count} results (no new results after 10 scrolls)")
                    break
            else:
                no_change_count = 0
            
            previous_count = current_count
            
            # Safety limit (very high)
            if scroll_count > 10000:
                print("\n⚠ Reached maximum scroll limit (10,000)")
                break
            
            # Every 100 scrolls, show a checkpoint
            if scroll_count % 100 == 0:
                print(f"\n[Checkpoint: {scroll_count} scrolls, {current_count} results]")
        
        print(f"Total scrolls: {scroll_count} | Final count: {current_count}")
    
    async def _wait_for_loading_complete(self, page):
        """Wait for loading spinners to completely disappear before continuing"""
        max_wait = 10  # Maximum 10 seconds to wait for loading
        wait_count = 0
        
        while wait_count < max_wait:
            # Check for common loading indicators
            loading_indicators = await page.query_selector_all('div[role="progressbar"], div.loading, div.spinner')
            
            if not loading_indicators:
                # No loading indicators found, we're good
                break
            
            # Still loading, wait a bit more
            await page.wait_for_timeout(1000)
            wait_count += 1
        
        # Extra small wait to ensure content has rendered
        await page.wait_for_timeout(500)
    
    async def _get_all_business_urls(self, page):
        """Extract all unique business URLs from the results - more thorough approach"""
        print("Extracting business URLs thoroughly...")
        
        # Try multiple selectors to catch all possible business listings
        selectors = [
            'a[href*="/maps/place/"]',  # Standard place links
            'a[data-value*="place"]',    # Alternative place links
            'div[role="feed"] a[href*="maps"]',  # Any maps link in feed
        ]
        
        all_links = []
        for selector in selectors:
            links = await page.query_selector_all(selector)
            all_links.extend(links)
        
        print(f"Found {len(all_links)} total links (including duplicates)")
        
        seen_urls = set()
        unique_urls = []
        
        for link in all_links:
            try:
                href = await link.get_attribute('href')
                if not href or '/maps/place/' not in href:
                    continue
                
                # Clean up the URL - remove query parameters for deduplication
                base_url = href.split('?')[0] if '?' in href else href
                
                if base_url in seen_urls:
                    continue
                
                # Build full URL
                if href.startswith('http'):
                    full_url = href
                else:
                    full_url = f"https://www.google.com{href}"
                
                seen_urls.add(base_url)
                unique_urls.append(full_url)
                
            except Exception as e:
                continue
        
        print(f"Extracted {len(unique_urls)} unique business URLs")
        return unique_urls
    
    async def _extract_business_data(self, page):
        """Extract business information from the detail page"""
        data = {
            'name': 'not_available',
            'rating': 'not_available',
            'reviews': 'not_available',
            'address': 'not_available',
            'phone': 'not_available',
            'website': 'not_available',
            'hours': 'not_available',
            'email': 'not_available',
            'social_media': 'not_available',
            'category': 'not_available',
            'plus_code': 'not_available'
        }
        
        try:
            # Wait for the main content to load
            await page.wait_for_selector('h1', timeout=10000)
            
            # Business name
            try:
                name_elem = await page.query_selector('h1.DUwDvf')
                if not name_elem:
                    name_elem = await page.query_selector('h1')
                if name_elem:
                    name_text = await name_elem.inner_text()
                    data['name'] = name_text.strip()
            except:
                pass
            
            # Rating
            try:
                rating_elem = await page.query_selector('div.F7nice span[aria-hidden="true"]')
                if rating_elem:
                    data['rating'] = await rating_elem.inner_text()
            except:
                pass
            
            # Number of reviews
            try:
                reviews_elem = await page.query_selector('div.F7nice span[aria-label*="reviews"]')
                if reviews_elem:
                    reviews_text = await reviews_elem.inner_text()
                    data['reviews'] = reviews_text.strip('()')
            except:
                pass
            
            # Address
            try:
                address_elem = await page.query_selector('button[data-item-id="address"] div.fontBodyMedium')
                if address_elem:
                    data['address'] = await address_elem.inner_text()
            except:
                pass
            
            # Phone number
            try:
                phone_elem = await page.query_selector('button[data-item-id*="phone:tel:"] div.fontBodyMedium')
                if phone_elem:
                    data['phone'] = await phone_elem.inner_text()
            except:
                pass
            
            # Website
            try:
                website_elem = await page.query_selector('a[data-item-id="authority"] div.fontBodyMedium')
                if website_elem:
                    data['website'] = await website_elem.inner_text()
            except:
                pass
            
            # Business hours
            try:
                hours_elem = await page.query_selector('button[data-item-id="oh"] div.fontBodyMedium')
                if hours_elem:
                    data['hours'] = await hours_elem.inner_text()
            except:
                pass
            
            # Try to find email
            try:
                email = await self._find_email(page)
                if email != 'not_available':
                    data['email'] = email
            except:
                pass
            
            # Social media links
            try:
                social = await self._find_social_media(page)
                if social != 'not_available':
                    data['social_media'] = social
            except:
                pass
            
            # Business type/category
            try:
                category_elem = await page.query_selector('button.DkEaL')
                if category_elem:
                    data['category'] = await category_elem.inner_text()
            except:
                pass
            
            # Plus Code / Location
            try:
                plus_code_elem = await page.query_selector('button[data-item-id="oloc"] div.fontBodyMedium')
                if plus_code_elem:
                    data['plus_code'] = await plus_code_elem.inner_text()
            except:
                pass
            
        except Exception as e:
            print(f"\n⚠ Error extracting data: {e}")
        
        return data
    
    async def _find_email(self, page):
        """Attempt to find email address"""
        try:
            content = await page.content()
            email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            emails = re.findall(email_pattern, content)
            # Filter out common non-business emails
            valid_emails = [e for e in emails if not any(x in e.lower() for x in ['google', 'schema.org', 'example.com'])]
            return valid_emails[0] if valid_emails else 'not_available'
        except:
            return 'not_available'
    
    async def _find_social_media(self, page):
        """Find social media links"""
        social_media = {}
        try:
            content = await page.content()
            
            # Common social media patterns
            platforms = {
                'facebook': r'(?:https?://)?(?:www\.)?facebook\.com/[\w\-\.]+',
                'instagram': r'(?:https?://)?(?:www\.)?instagram\.com/[\w\-\.]+',
                'twitter': r'(?:https?://)?(?:www\.)?(?:twitter|x)\.com/[\w\-\.]+',
                'linkedin': r'(?:https?://)?(?:www\.)?linkedin\.com/(?:company|in)/[\w\-\.]+',
                'youtube': r'(?:https?://)?(?:www\.)?youtube\.com/[\w\-\.]+',
            }
            
            for platform, pattern in platforms.items():
                matches = re.findall(pattern, content)
                if matches:
                    # Get the first unique match
                    social_media[platform] = matches[0]
            
            return json.dumps(social_media) if social_media else 'not_available'
            
        except Exception as e:
            return 'not_available'
    
    def _sanitize_filename(self, query):
        """Convert search query to a clean filename"""
        # Convert to lowercase and replace spaces with underscores
        filename = query.lower().replace(' ', '_')
        # Remove any characters that aren't alphanumeric, underscore, or hyphen
        filename = re.sub(r'[^\w\-]', '', filename)
        return filename
    
    def save_to_csv(self, filename=None, search_query=None):
        """Save results to CSV file"""
        if not self.results:
            print("No results to save")
            return
        
        if not filename:
            if search_query:
                base_name = self._sanitize_filename(search_query)
                filename = f"{base_name}.csv"
            else:
                filename = f"gmaps_scrape_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        keys = self.results[0].keys()
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(self.results)
        
        print(f"\n✓ Results saved to {filename}")
        return filename
    
    def save_to_json(self, filename=None, search_query=None):
        """Save results to JSON file"""
        if not self.results:
            print("No results to save")
            return
        
        if not filename:
            if search_query:
                base_name = self._sanitize_filename(search_query)
                filename = f"{base_name}.json"
            else:
                filename = f"gmaps_scrape_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False)
        
        print(f"✓ Results saved to {filename}")
        return filename
    
    def save_to_excel(self, filename=None, search_query=None):
        """Save results to Excel file with formatting"""
        if not self.results:
            print("No results to save")
            return
        
        if not filename:
            if search_query:
                base_name = self._sanitize_filename(search_query)
                filename = f"{base_name}.xlsx"
            else:
                filename = f"gmaps_scrape_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Business Data"
        
        # Add headers
        headers = list(self.results[0].keys())
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data
        for result in self.results:
            ws.append(list(result.values()))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        print(f"✓ Results saved to {filename}")
        return filename

queries = ["add all the queries to run on google maps to search google businesses", "like Freight brokers in Tracy"
]

async def main():
    scraper = GoogleMapsScraper()
    
    total_queries = len(queries)
    max_results = None  # Set to None to scrape ALL results until end of list or define a limit
    
    print("=" * 60)
    print("Google Maps Business Scraper - ROBUST VERSION")
    print(f"Total Queries to Process: {total_queries}")
    print("=" * 60)
    print()
    
    # Loop through each query in the queries array
    for query_index, search_query in enumerate(queries, 1):
        # Clear results before each query to avoid mixing data
        scraper.results = []
        
        print("\n" + "=" * 60)
        print(f"QUERY {query_index}/{total_queries}")
        print("=" * 60)
        print(f"Search Query: {search_query}")
        print(f"Max Results: {'UNLIMITED (scrape until end)' if max_results is None else max_results}")
        print("=" * 60)
        
        try:
            # Scrape businesses for this query
            results = await scraper.scrape_businesses(search_query, max_results)
            
            print(f"\n{'=' * 60}")
            print(f"✓ Query {query_index}/{total_queries} Complete: {len(results)} businesses scraped")
            print(f"{'=' * 60}")
            
            # Save results in all formats for this query
            if results:
                scraper.save_to_json(search_query=search_query)
                scraper.save_to_csv(search_query=search_query)
                scraper.save_to_excel(search_query=search_query)
                
                # Display summary for this query
                print("\n" + "-" * 60)
                print(f"RESULTS PREVIEW FOR: {search_query}")
                print("-" * 60)
                for i, business in enumerate(results[:5], 1):
                    print(f"\n{i}. {business.get('name', 'not_available')}")
                    print(f"   Phone: {business.get('phone', 'not_available')}")
                    print(f"   Email: {business.get('email', 'not_available')}")
                    print(f"   Website: {business.get('website', 'not_available')}")
                    print(f"   Address: {business.get('address', 'not_available')}")
                    print(f"   Rating: {business.get('rating', 'not_available')} ({business.get('reviews', 'not_available')} reviews)")
            else:
                print(f"⚠ No results found for query: {search_query}")
        
        except Exception as e:
            print(f"\n✗ Error processing query '{search_query}': {e}")
            print("Continuing with next query...")
            continue
        
        # Add separator between queries
        if query_index < total_queries:
            print(f"\n\n{'#' * 60}")
            print(f"Moving to next query...")
            print(f"{'#' * 60}\n")
            await asyncio.sleep(2)  # Small delay between queries
    
    # Final summary
    print("\n" + "=" * 60)
    print("ALL QUERIES COMPLETED!")
    print(f"Processed {total_queries} queries")
    print("=" * 60)


if __name__ == "__main__":
    # Install required packages first:
    # pip install playwright openpyxl
    # playwright install chromium
    
    asyncio.run(main())
